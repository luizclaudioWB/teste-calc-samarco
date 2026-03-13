#!/usr/bin/env python3
"""
Extract formulas and data from Samarco Excel workbook.
Targets: Gross-Up, Distribuição de Carga, Classe de Custo, Centro de Custos, Resumo Geral
"""

import openpyxl
import csv
import os

XLSX = '/home/wise/Vaults/notas-rascunhos/TesteMotorCalculo/SAMARCO-PLANILHA-REPLY.xlsx'

TARGET_KEYWORDS = ["gross", "distribui", "classe", "centro", "resumo"]

def col_letter(n):
    """Convert 1-based column index to Excel column letter."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result

def matches_target(name):
    return any(kw in name.lower() for kw in TARGET_KEYWORDS)

# ── Load both workbooks ────────────────────────────────────────────────────────
wb_vals = openpyxl.load_workbook(XLSX, data_only=True)
wb_fmls = openpyxl.load_workbook(XLSX, data_only=False)

# ── 1. All sheet names ─────────────────────────────────────────────────────────
print("=" * 80)
print("ALL SHEET NAMES")
print("=" * 80)
for i, name in enumerate(wb_vals.sheetnames, 1):
    marker = " <-- TARGET" if matches_target(name) else ""
    print(f"  {i:02d}. {name}{marker}")
print()

# ── 2. Per-target sheet: full formula + value dump ─────────────────────────────
target_sheets = [n for n in wb_vals.sheetnames if matches_target(n)]

for sheet_name in target_sheets:
    ws_v = wb_vals[sheet_name]
    ws_f = wb_fmls[sheet_name]

    max_row = ws_v.max_row
    max_col = ws_v.max_column

    print()
    print("=" * 80)
    print(f"SHEET: {sheet_name}")
    print(f"  Dimensions: {max_row} rows x {max_col} cols ({col_letter(max_col)})")
    print("=" * 80)

    # Header row
    print("\n  --- HEADER ROW (row 1) ---")
    for c in range(1, max_col + 1):
        v = ws_v.cell(row=1, column=c).value
        if v is not None:
            print(f"    {col_letter(c)}1 = {repr(v)}")

    # All data rows
    print("\n  --- ALL ROWS ---")
    for r in range(1, max_row + 1):
        # Skip entirely blank rows
        row_empty = True
        for c in range(1, max_col + 1):
            if ws_v.cell(row=r, column=c).value is not None:
                row_empty = False
                break
            if ws_f.cell(row=r, column=c).value is not None:
                row_empty = False
                break
        if row_empty:
            continue

        label = ws_v.cell(row=r, column=1).value
        print(f"\n  ROW {r:03d} | A{r} = {repr(label)}")

        # Formulas: columns B through D (first 3 months)
        fml_parts = []
        for c in range(2, min(5, max_col + 1)):
            fv = ws_f.cell(row=r, column=c).value
            if fv is not None:
                fml_parts.append(f"{col_letter(c)}{r}={repr(fv)}")
        if fml_parts:
            print(f"    FORMULAS (B-D): {' | '.join(fml_parts)}")

        # Values: all columns B through max
        val_parts = []
        for c in range(2, max_col + 1):
            vv = ws_v.cell(row=r, column=c).value
            if vv is not None:
                val_parts.append(f"{col_letter(c)}{r}={repr(vv)}")
        if val_parts:
            print(f"    VALUES        : {' | '.join(val_parts)}")

    print(f"\n  [END: {sheet_name}]")

# ── 3. Extended formula dump — all formulas through col N ─────────────────────
print()
print("=" * 80)
print("EXTENDED FORMULA DUMP — ALL formula cells cols B-N per target sheet")
print("=" * 80)

for sheet_name in target_sheets:
    ws_v = wb_vals[sheet_name]
    ws_f = wb_fmls[sheet_name]

    print(f"\n=== {sheet_name} ===")
    max_row = ws_f.max_row
    max_col = ws_f.max_column
    end_col = min(14, max_col)  # up to col N

    for r in range(1, max_row + 1):
        label = ws_v.cell(row=r, column=1).value
        found = []
        for c in range(2, end_col + 1):
            fv = ws_f.cell(row=r, column=c).value
            if fv is not None and str(fv).startswith("="):
                found.append(f"  {col_letter(c)}{r} = {repr(fv)}")
        if found:
            print(f"\n  ROW {r:03d} [{repr(label)}]:")
            for line in found:
                print(line)

print()
print("DONE.")
